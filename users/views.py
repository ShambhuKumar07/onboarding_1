from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth import logout
from .forms import CustomAuthenticationForm 
from compliance.models import Employee
from compensation.models import EmployeeCTC
import openpyxl
from django.http import HttpResponse
from users.models import Profile
from applicant.models import Applicant
from django.utils.dateparse import parse_date
from django.db.models import Q
from io import BytesIO
import openpyxl
from .models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            login(request, form.user_cache)
            try:
                if request.user.applicant:  # Trying to access the related applicant
                    return redirect('applicant:dashboard')
            except request.user._meta.model.applicant.RelatedObjectDoesNotExist:
                return redirect('core:main_page')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'users/login.html', {'form': form})

def logout_view(request):
    logout(request)

    return redirect('users:login_view')

########################### Modifying code Personal Information for columns wise

def export_profile_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references, compliance_info, compensation_info):
    import openpyxl
    from openpyxl.styles import Alignment
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Profile Details"

    # Define headers for the personal details section
    personal_headers = [
        "Employee Code", "First Name", "Middle Name", "Last Name",
        "Designation", "Date of Joining", "End Date", "Email",
        "Phone Number", "Gender", "Date of Birth", "Present Address",
        "Permanent Address", "Personal Email", "PAN Number", "Marital Status"
    ]
    personal_details = [
        profile.employee_code, profile.first_name, profile.middle_name,
        profile.last_name, profile.designation, profile.date_of_joining,
        profile.end_date, profile.email, profile.phone_number,
        profile.get_gender_display(), profile.date_of_birth,
        profile.present_address, profile.permanent_address,
        profile.personal_email, profile.pan_number, profile.marital_status
    ]
    # Write headers and personal details in column-wise format
    for col, header in enumerate(personal_headers, start=1):
        sheet.cell(row=1, column=col, value=header)  # Header row
        sheet.cell(row=2, column=col, value=personal_details[col - 1])  # Data row

    # Adjust alignment for readability
    for col in range(1, len(personal_headers) + 1):
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        sheet.cell(row=2, column=col).alignment = Alignment(horizontal='center')

    # Add additional sections for family members, education, etc.
    row_offset = 4

    # Family Members Section
    sheet.cell(row=row_offset, column=1, value="Family Members")
    row_offset += 1
    sheet.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
    for member in family_members:
        sheet.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])
    row_offset += len(family_members) + 2

    # Educational Qualifications
    sheet.cell(row=row_offset, column=1, value="Educational Qualifications")
    row_offset += 1
    sheet.append(["Examination Passed", "Year of Passing", "School/College", "Division"])
    for qualification in educational_qualifications:
        sheet.append([
            qualification.examination_passed,
            qualification.year_of_passing,
            qualification.school_or_college,
            qualification.division,
        ])
    row_offset += len(educational_qualifications) + 2

    # Employment Records
    sheet.cell(row=row_offset, column=1, value="Employment Records")
    row_offset += 1
    sheet.append(["Organization", "Designation", "Joining Date", "Leaving Date"])
    for record in employment_records:
        sheet.append([
            record.organization, record.designation,
            record.joining_date, record.leaving_date
        ])
    row_offset += len(employment_records) + 2

    # Language Proficiencies
    sheet.cell(row=row_offset, column=1, value="Languages Known")
    row_offset += 1
    sheet.append(["Language", "Speak", "Read", "Write"])
    for language in language_proficiencies:
        sheet.append([
            language.language,
            language.get_speak_display(),
            language.get_read_display(),
            language.get_write_display()
        ])
    row_offset += len(language_proficiencies) + 2

    # References
    sheet.cell(row=row_offset, column=1, value="References")
    row_offset += 1
    sheet.append(["Name", "Occupation", "Phone Number", "Email"])
    for reference in references:
        sheet.append([
            reference.name, reference.occupation,
            reference.phone_number, reference.email
        ])
    row_offset += len(references) + 2

    # Compliance Information
    sheet.cell(row=row_offset, column=1, value="Compliance Information")
    row_offset += 1
    if compliance_info:
        sheet.append(["UAN Number", compliance_info.uan_number])
        sheet.append(["PAN Number", compliance_info.pan_number])
        sheet.append(["ESIC Number", compliance_info.esic_number])
    else:
        sheet.append(["No compliance data available"])
    row_offset += 3

    # Compensation Information
    sheet.cell(row=row_offset, column=1, value="Compensation Information")
    row_offset += 1
    if compensation_info:
        sheet.append(["Component", "Amount"])
        salary = compensation_info.salary
        contribution = compensation_info.contribution
        deduction = compensation_info.deduction
        sheet.append(["Basic Salary", salary.basic])
        sheet.append(["HRA", salary.hra])
        sheet.append(["Special Allowance", salary.special_allowance])
        sheet.append(["Gross Salary", salary.gross_salary()])
        sheet.append(["Total Contributions", contribution.emp_contribution()])
        sheet.append(["Total Deductions", deduction.emp_deduction()])
        sheet.append(["Net Salary (CTC)", compensation_info.calculate_ctc()])
    else:
        sheet.append(["No compensation data available"])

    return workbook

################################ update views for compliance data
def profile_detail_view(request, profile_id):
    # Get the specific profile by its ID
    profile = get_object_or_404(Profile, id=profile_id)
    all_profiles = Profile.objects.all()

    # Access related FamilyMember, EducationalQualification, etc.
    family_members = profile.family_members.all()
    educational_qualifications = profile.educational_qualifications.all()
    employment_records = profile.employment_records.all()
    language_proficiencies = profile.languages.all()
    references = profile.references.all()

    # Fetch the compliance data through the Applicant linked to this Profile
    applicant = getattr(profile.user, 'applicant', None)  # Access Applicant via the CustomUser relationship
    compliance_info = Employee.objects.filter(applicant=applicant).first() if applicant else None
    compensation_info = EmployeeCTC.objects.filter(applicant=applicant).first()
    # compensation_info = EmployeeCTC.objects.filter(profile=profile).first()

    # Export to Excel if requested
    if request.GET.get('export') == 'excel':
        workbook = export_profile_to_excel(
            profile,
            family_members,
            educational_qualifications,
            employment_records,
            language_proficiencies,
            references,
            compliance_info,
            compensation_info  # Add these two arguments
        )
    # Prepare the response
        response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Profile_{profile.employee_code}.xlsx'
        workbook.save(response)
        return response
    # Pass the data to the template if not exporting
    context = {
        'profile': profile,
        'all_profiles': all_profiles,
        'family_members': family_members,
        'educational_qualifications': educational_qualifications,
        'employment_records': employment_records,
        'language_proficiencies': language_proficiencies,
        'references': references,
        'compliance_info': compliance_info,
        'compensation_info': compensation_info,
    }

    return render(request, 'profile_detail.html', context)

def hiring_report(request):
    from_date = request.GET.get('from_date')
    date_of_joining = request.GET.get('date_of_joining')
    
    profiles = Profile.objects.all()
    
    # Apply the filter based on 'date_of_joining' range
    if from_date and date_of_joining:
        profiles = profiles.filter(
            date_of_joining__range=(parse_date(from_date), parse_date(date_of_joining))
        )
    elif from_date:
        profiles = profiles.filter(date_of_joining__gte=parse_date(from_date))
    elif date_of_joining:
        profiles = profiles.filter(date_of_joining__lte=parse_date(date_of_joining))
    
    return render(request, 'users/hiring_report.html', {'profiles': profiles, 'from_date': from_date, 'date_of_joining': date_of_joining})


def attrition_report(request):
    from_date = request.GET.get('from_date')
    end_date = request.GET.get('end_date')
    
    profiles = Profile.objects.filter(end_date__isnull=False)  # Only include profiles with an end_date

    # Apply the filter based only on the 'end_date' range
    if from_date and end_date:
        profiles = profiles.filter(
            end_date__range=(parse_date(from_date), parse_date(end_date))
        )
    elif from_date:
        profiles = profiles.filter(end_date__gte=parse_date(from_date))
    elif end_date:
        profiles = profiles.filter(end_date__lte=parse_date(end_date))
    
    return render(request, 'users/attrition_report.html', {'profiles': profiles, 'from_date': from_date, 'end_date': end_date})


########################### export all applicant details
########################## Add compensation and complianc  #################

# from django.http import HttpResponse
# from openpyxl import Workbook
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference


# def export_applicants_to_excel(request):
#     # Create a new Excel workbook and active sheet
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     # Define the headers for the Excel sheet
#     headers = [
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date", 
#         "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email", 
#         "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address", 
#         "Marital Status", "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number", 
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",
#         "Family Members", "Educational Qualifications", "Employment Records", "Language Proficiencies", 
#         "References", "UAN Number", "PAN Number (Compliance)", "ESIC Number", "Gross Salary", 
#         "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     # Write headers to the first row of the sheet
#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=1, column=col_idx, value=header)

#     # Fetch applicants and write their data row by row
#     applicants = Applicant.objects.select_related('profile', 'user')
#     for row_idx, applicant in enumerate(applicants, start=2):
#         profile = applicant.profile

#         # Write profile data
#         sheet.cell(row=row_idx, column=1, value=profile.fullname)
#         sheet.cell(row=row_idx, column=2, value=profile.employee_code)
#         sheet.cell(row=row_idx, column=3, value=profile.designation)
#         sheet.cell(row=row_idx, column=4, value=profile.date_of_joining)
#         sheet.cell(row=row_idx, column=5, value=profile.end_date)
#         sheet.cell(row=row_idx, column=6, value=profile.get_gender_display())
#         sheet.cell(row=row_idx, column=7, value=profile.date_of_birth)
#         sheet.cell(row=row_idx, column=8, value=profile.age)
#         sheet.cell(row=row_idx, column=9, value=profile.email)
#         sheet.cell(row=row_idx, column=10, value=profile.phone_number)
#         sheet.cell(row=row_idx, column=11, value=profile.personal_email)
#         sheet.cell(row=row_idx, column=12, value=profile.father_name)
#         sheet.cell(row=row_idx, column=13, value=profile.mother_name)
#         sheet.cell(row=row_idx, column=14, value=profile.blood_group)
#         sheet.cell(row=row_idx, column=15, value=profile.present_address)
#         sheet.cell(row=row_idx, column=16, value=profile.permanent_address)
#         sheet.cell(row=row_idx, column=17, value=profile.marital_status)
#         sheet.cell(row=row_idx, column=18, value=profile.pan_number)
#         sheet.cell(row=row_idx, column=19, value=profile.bank_name)
#         sheet.cell(row=row_idx, column=20, value=profile.ifsc_code)
#         sheet.cell(row=row_idx, column=21, value=profile.bank_account_number)
#         sheet.cell(row=row_idx, column=22, value=profile.emergency_contact_name)
#         sheet.cell(row=row_idx, column=23, value=profile.emergency_contact_relation)
#         sheet.cell(row=row_idx, column=24, value=profile.emergency_contact_number)

#         # Write related data as concatenated strings
#         # Family Members
#         family_members = ", ".join([f"{fm.name} ({fm.relation})" for fm in profile.family_members.all()])
#         sheet.cell(row=row_idx, column=25, value=family_members)

#         # Educational Qualifications
#         educational_qualifications = ", ".join([
#             f"{eq.examination_passed} ({eq.year_of_passing})" for eq in profile.educational_qualifications.all()
#         ])
#         sheet.cell(row=row_idx, column=26, value=educational_qualifications)

#         # Employment Records
#         employment_records = ", ".join([
#             f"{er.organization} ({er.designation})" for er in profile.employment_records.all()
#         ])
#         sheet.cell(row=row_idx, column=27, value=employment_records)

#         # Language Proficiencies
#         languages = ", ".join([
#             f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})" 
#             for lp in profile.languages.all()
#         ])
#         sheet.cell(row=row_idx, column=28, value=languages)

#         # References
#         references = ", ".join([
#             f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all()
#         ])
#         sheet.cell(row=row_idx, column=29, value=references)

#         # Add Compliance details
#         try:
#             compliance = applicant.employee
#             sheet.cell(row=row_idx, column=30, value=compliance.uan_number)
#             sheet.cell(row=row_idx, column=31, value=compliance.pan_number)
#             sheet.cell(row=row_idx, column=32, value=compliance.esic_number)
#         except AttributeError:
#             # If no compliance record exists
#             sheet.cell(row=row_idx, column=30, value="N/A")
#             sheet.cell(row=row_idx, column=31, value="N/A")
#             sheet.cell(row=row_idx, column=32, value="N/A")

#         # Add Compensation details
#         try:
#             compensation = applicant.employeectc
#             sheet.cell(row=row_idx, column=33, value=compensation.salary.gross_salary())
#             sheet.cell(row=row_idx, column=34, value=compensation.contribution.emp_contribution())
#             sheet.cell(row=row_idx, column=35, value=compensation.deduction.emp_deduction())
#             sheet.cell(row=row_idx, column=36, value=compensation.calculate_ctc())
#         except AttributeError:
#             # If no compensation record exists
#             sheet.cell(row=row_idx, column=33, value="N/A")
#             sheet.cell(row=row_idx, column=34, value="N/A")
#             sheet.cell(row=row_idx, column=35, value="N/A")
#             sheet.cell(row=row_idx, column=36, value="N/A")

#     # Set the HTTP response with the Excel file
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response


   
from django.http import HttpResponse
from openpyxl import Workbook
from applicant.models import Applicant
from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference


# def export_applicants_to_excel(request):
#     # Create a new Excel workbook and active sheet
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     # Define the headers for the Excel sheet
#     headers = [
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date",
#         "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email",
#         "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address",
#         "Marital Status", "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",
#         "Family Members", "Educational Qualifications", "Employment Records", "Language Proficiencies",
#         "References", "UAN Number", "PAN Number (Compliance)", "ESIC Number", "Gross Salary",
#         "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     # Write headers to the first row of the sheet
#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=1, column=col_idx, value=header)

#     # Fetch applicants and related data
#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members',
#         'profile__educational_qualifications',
#         'profile__employment_records',
#         'profile__languages',
#         'profile__references',
#     )

#     for row_idx, applicant in enumerate(applicants, start=2):
#         profile = applicant.profile

#         # Write profile data
#         sheet.cell(row=row_idx, column=1, value=profile.fullname)
#         sheet.cell(row=row_idx, column=2, value=profile.employee_code)
#         sheet.cell(row=row_idx, column=3, value=profile.designation)
#         sheet.cell(row=row_idx, column=4, value=profile.date_of_joining)
#         sheet.cell(row=row_idx, column=5, value=profile.end_date)
#         sheet.cell(row=row_idx, column=6, value=profile.get_gender_display())
#         sheet.cell(row=row_idx, column=7, value=profile.date_of_birth)
#         sheet.cell(row=row_idx, column=8, value=profile.age)
#         sheet.cell(row=row_idx, column=9, value=profile.email)
#         sheet.cell(row=row_idx, column=10, value=profile.phone_number)
#         sheet.cell(row=row_idx, column=11, value=profile.personal_email)
#         sheet.cell(row=row_idx, column=12, value=profile.father_name)
#         sheet.cell(row=row_idx, column=13, value=profile.mother_name)
#         sheet.cell(row=row_idx, column=14, value=profile.blood_group)
#         sheet.cell(row=row_idx, column=15, value=profile.present_address)
#         sheet.cell(row=row_idx, column=16, value=profile.permanent_address)
#         sheet.cell(row=row_idx, column=17, value=profile.marital_status)
#         sheet.cell(row=row_idx, column=18, value=profile.pan_number)
#         sheet.cell(row=row_idx, column=19, value=profile.bank_name)
#         sheet.cell(row=row_idx, column=20, value=profile.ifsc_code)
#         sheet.cell(row=row_idx, column=21, value=profile.bank_account_number)
#         sheet.cell(row=row_idx, column=22, value=profile.emergency_contact_name)
#         sheet.cell(row=row_idx, column=23, value=profile.emergency_contact_relation)
#         sheet.cell(row=row_idx, column=24, value=profile.emergency_contact_number)

#         # Write related data
#         family_members = ", ".join([f"{fm.name} ({fm.relation})" for fm in profile.family_members.all()])
#         sheet.cell(row=row_idx, column=25, value=family_members)

#         educational_qualifications = ", ".join([
#             f"{eq.examination_passed} ({eq.year_of_passing})" for eq in profile.educational_qualifications.all()
#         ])
#         sheet.cell(row=row_idx, column=26, value=educational_qualifications)

#         employment_records = ", ".join([
#             f"{er.organization} ({er.designation})" for er in profile.employment_records.all()
#         ])
#         sheet.cell(row=row_idx, column=27, value=employment_records)

#         languages = ", ".join([
#             f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})"
#             for lp in profile.languages.all()
#         ])
#         sheet.cell(row=row_idx, column=28, value=languages)

#         references = ", ".join([
#             f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all()
#         ])
#         sheet.cell(row=row_idx, column=29, value=references)

#         # Compliance details
#         try:
#             compliance = applicant.employee
#             sheet.cell(row=row_idx, column=30, value=compliance.uan_number)
#             sheet.cell(row=row_idx, column=31, value=compliance.pan_number)
#             sheet.cell(row=row_idx, column=32, value=compliance.esic_number)
#         except AttributeError:
#             sheet.cell(row=row_idx, column=30, value="N/A")
#             sheet.cell(row=row_idx, column=31, value="N/A")
#             sheet.cell(row=row_idx, column=32, value="N/A")

#         # Compensation details
#         try:
#             compensation = EmployeeCTC.objects.get(applicant=applicant)
#             sheet.cell(row=row_idx, column=33, value=compensation.salary.gross_salary())
#             sheet.cell(row=row_idx, column=34, value=compensation.contribution.emp_contribution())
#             sheet.cell(row=row_idx, column=35, value=compensation.deduction.emp_deduction())
#             sheet.cell(row=row_idx, column=36, value=compensation.calculate_ctc())
#         except EmployeeCTC.DoesNotExist:
#             sheet.cell(row=row_idx, column=33, value="N/A")
#             sheet.cell(row=row_idx, column=34, value="N/A")
#             sheet.cell(row=row_idx, column=35, value="N/A")
#             sheet.cell(row=row_idx, column=36, value="N/A")

#     # Set the HTTP response with the Excel file
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response



############################### personal Information Header changes ####################

# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference


# def export_applicants_to_excel(request):
#     # Create a new Excel workbook and active sheet
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     # Merge cells for the "Personal Information" header
#     sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
#     sheet.cell(row=1, column=1, value="Personal Information")
#     sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

#     # Define headers for personal information and other sections
#     headers = [
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date",  # Personal Information
#         "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email",
#         "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address",
#         "Marital Status", "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",
#         "Family Members", "Educational Qualifications", "Employment Records", "Language Proficiencies",
#         "References", "UAN Number", "PAN Number (Compliance)", "ESIC Number", "Gross Salary",
#         "Total Contribution", "Total Deductions", "Net CTC"
#     ]


#     # # Define headers grouped by sections
#     # headers = [
#     # # Personal Information
#     # "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date",
#     # "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email",
#     # "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address",
#     # "Marital Status",

#     # # Banking Information
#     # "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",

#     # # Emergency Contact Information
#     # "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",

#     # # Family, Education, Employment, and Language Details
#     # "Family Members", "Educational Qualifications", "Employment Records", "Language Proficiencies",

#     # # References
#     # "References",

#     # # Compliance Information
#     # "UAN Number", "PAN Number (Compliance)", "ESIC Number",

#     # # Compensation Details
#     # "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
#     # ]


#     # Write headers for the second row (starting from Applicant Name)
#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)

#     # Fetch applicants and related data
#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members',
#         'profile__educational_qualifications',
#         'profile__employment_records',
#         'profile__languages',
#         'profile__references',
#     )

#     for row_idx, applicant in enumerate(applicants, start=3):
#         profile = applicant.profile

#         # Write profile data
#         sheet.cell(row=row_idx, column=1, value=profile.fullname)
#         sheet.cell(row=row_idx, column=2, value=profile.employee_code)
#         sheet.cell(row=row_idx, column=3, value=profile.designation)
#         sheet.cell(row=row_idx, column=4, value=profile.date_of_joining)
#         sheet.cell(row=row_idx, column=5, value=profile.end_date)
#         sheet.cell(row=row_idx, column=6, value=profile.get_gender_display())
#         sheet.cell(row=row_idx, column=7, value=profile.date_of_birth)
#         sheet.cell(row=row_idx, column=8, value=profile.age)
#         sheet.cell(row=row_idx, column=9, value=profile.email)
#         sheet.cell(row=row_idx, column=10, value=profile.phone_number)
#         sheet.cell(row=row_idx, column=11, value=profile.personal_email)
#         sheet.cell(row=row_idx, column=12, value=profile.father_name)
#         sheet.cell(row=row_idx, column=13, value=profile.mother_name)
#         sheet.cell(row=row_idx, column=14, value=profile.blood_group)
#         sheet.cell(row=row_idx, column=15, value=profile.present_address)
#         sheet.cell(row=row_idx, column=16, value=profile.permanent_address)
#         sheet.cell(row=row_idx, column=17, value=profile.marital_status)
#         sheet.cell(row=row_idx, column=18, value=profile.pan_number)
#         sheet.cell(row=row_idx, column=19, value=profile.bank_name)
#         sheet.cell(row=row_idx, column=20, value=profile.ifsc_code)
#         sheet.cell(row=row_idx, column=21, value=profile.bank_account_number)
#         sheet.cell(row=row_idx, column=22, value=profile.emergency_contact_name)
#         sheet.cell(row=row_idx, column=23, value=profile.emergency_contact_relation)
#         sheet.cell(row=row_idx, column=24, value=profile.emergency_contact_number)

#         # Write related data
#         family_members = ", ".join([f"{fm.name} ({fm.relation})" for fm in profile.family_members.all()])
#         sheet.cell(row=row_idx, column=25, value=family_members)

#         educational_qualifications = ", ".join([
#             f"{eq.examination_passed} ({eq.year_of_passing})" for eq in profile.educational_qualifications.all()
#         ])
#         sheet.cell(row=row_idx, column=26, value=educational_qualifications)

#         employment_records = ", ".join([
#             f"{er.organization} ({er.designation})" for er in profile.employment_records.all()
#         ])
#         sheet.cell(row=row_idx, column=27, value=employment_records)

#         languages = ", ".join([
#             f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})"
#             for lp in profile.languages.all()
#         ])
#         sheet.cell(row=row_idx, column=28, value=languages)

#         references = ", ".join([
#             f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all()
#         ])
#         sheet.cell(row=row_idx, column=29, value=references)

#         # Compliance details
#         try:
#             compliance = applicant.employee
#             sheet.cell(row=row_idx, column=30, value=compliance.uan_number)
#             sheet.cell(row=row_idx, column=31, value=compliance.pan_number)
#             sheet.cell(row=row_idx, column=32, value=compliance.esic_number)
#         except AttributeError:
#             sheet.cell(row=row_idx, column=30, value="N/A")
#             sheet.cell(row=row_idx, column=31, value="N/A")
#             sheet.cell(row=row_idx, column=32, value="N/A")

#         # Compensation details
#         try:
#             compensation = EmployeeCTC.objects.get(applicant=applicant)
#             sheet.cell(row=row_idx, column=33, value=compensation.salary.gross_salary())
#             sheet.cell(row=row_idx, column=34, value=compensation.contribution.emp_contribution())
#             sheet.cell(row=row_idx, column=35, value=compensation.deduction.emp_deduction())
#             sheet.cell(row=row_idx, column=36, value=compensation.calculate_ctc())
#         except EmployeeCTC.DoesNotExist:
#             sheet.cell(row=row_idx, column=33, value="N/A")
#             sheet.cell(row=row_idx, column=34, value="N/A")
#             sheet.cell(row=row_idx, column=35, value="N/A")
#             sheet.cell(row=row_idx, column=36, value="N/A")

#     # Set the HTTP response with the Excel file
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response


############################


# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference



# def export_applicants_to_excel(request):
#     # Create a new Excel workbook and active sheet
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     # Merge cells for the "Personal Information" header
#     sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
#     sheet.cell(row=1, column=1, value="Personal Information")
#     sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

#     # Merge cells for the "Banking Information" header
#     sheet.merge_cells(start_row=1, start_column=18, end_row=1, end_column=21)
#     sheet.cell(row=1, column=18, value="Banking Information")
#     sheet.cell(row=1, column=18).alignment = Alignment(horizontal="center", vertical="center")

#     # Merge cells for the "Emergency Contact Information" header
#     sheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=24)
#     sheet.cell(row=1, column=22, value="Emergency Contact Information")
#     sheet.cell(row=1, column=22).alignment = Alignment(horizontal="center", vertical="center")


#     # Merge cells for the "Emergency Contact Information" header
#     sheet.merge_cells(start_row=1, start_column=25, end_row=1, end_column=28)
#     sheet.cell(row=1, column=22, value=" Famaly Memeber")
#     sheet.cell(row=1, column=22).alignment = Alignment(horizontal="center", vertical="center")



#     # Merge cells for other sections if needed...

#     # Define headers for the second row
#     headers = [
#         # Personal Information
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date",
#         "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email",
#         "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address",
#         "Marital Status",
         

#         # Banking Information
#         "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",

#         # Emergency Contact Information
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",

#         # Family
#         "Relation", "Name", "DOB", "Sex",

#         # Family, Education, Employment, and Language Details
#         "Educational Qualifications", "Employment Records", "Language Proficiencies",

#         # References
#         "References",

#         # Compliance Information
#         "UAN Number", "PAN Number (Compliance)", "ESIC Number",

#         # Compensation Details
#         "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     # Write headers for the second row
#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)

#     # Fetch applicants and related data
#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members',
#         'profile__educational_qualifications',
#         'profile__employment_records',
#         'profile__languages',
#         'profile__references',
#     )

#     for row_idx, applicant in enumerate(applicants, start=3):
#         profile = applicant.profile
#         # family_members = profile.family_members.all()

#         # Write profile data
#         sheet.cell(row=row_idx, column=1, value=profile.fullname)
#         sheet.cell(row=row_idx, column=2, value=profile.employee_code)
#         sheet.cell(row=row_idx, column=3, value=profile.designation)
#         sheet.cell(row=row_idx, column=4, value=profile.date_of_joining)
#         sheet.cell(row=row_idx, column=5, value=profile.end_date)
#         sheet.cell(row=row_idx, column=6, value=profile.get_gender_display())
#         sheet.cell(row=row_idx, column=7, value=profile.date_of_birth)
#         sheet.cell(row=row_idx, column=8, value=profile.age)
#         sheet.cell(row=row_idx, column=9, value=profile.email)
#         sheet.cell(row=row_idx, column=10, value=profile.phone_number)
#         sheet.cell(row=row_idx, column=11, value=profile.personal_email)
#         sheet.cell(row=row_idx, column=12, value=profile.father_name)
#         sheet.cell(row=row_idx, column=13, value=profile.mother_name)
#         sheet.cell(row=row_idx, column=14, value=profile.blood_group)
#         sheet.cell(row=row_idx, column=15, value=profile.present_address)
#         sheet.cell(row=row_idx, column=16, value=profile.permanent_address)
#         sheet.cell(row=row_idx, column=17, value=profile.marital_status)
#         sheet.cell(row=row_idx, column=18, value=profile.pan_number)
#         sheet.cell(row=row_idx, column=19, value=profile.bank_name)
#         sheet.cell(row=row_idx, column=20, value=profile.ifsc_code)
#         sheet.cell(row=row_idx, column=21, value=profile.bank_account_number)
#         sheet.cell(row=row_idx, column=22, value=profile.emergency_contact_name)
#         sheet.cell(row=row_idx, column=23, value=profile.emergency_contact_relation)
#         sheet.cell(row=row_idx, column=24, value=profile.emergency_contact_number)
#         # sheet.cell(row=row_idx, column=25, value=FamilyMember.name)
#         # sheet.cell(row=row_idx, column=26, value=FamilyMember.relation)
#         # sheet.cell(row=row_idx, column=27, value=FamilyMember.date_of_birth)
#         # sheet.cell(row=row_idx, column=28, value=FamilyMember.sex)
#         # sheet.cell(row=row_idx, column=29, value=FamilyMember.age)


# ################### Family Memeber attribute   ##################



#                 # Write related data
#         # family_members = ", ".join([f"{fm.name} ({fm.relation})" for fm in profile.family_members.all()])
#         # sheet.cell(row=row_idx, column=25, value=family_members)

#         # Write family member details

#         # Write family member details
#         family_member = profile.family_members.first()  # Example: Get the first family member
#         if family_member:
#             sheet.cell(row=row_idx, column=25, value=family_member.name)
#             sheet.cell(row=row_idx, column=26, value=family_member.relation)
#             sheet.cell(row=row_idx, column=27, value=family_member.date_of_birth)
#             sheet.cell(row=row_idx, column=28, value=family_member.sex)
#             sheet.cell(row=row_idx, column=29, value=family_member.age)
#         else:
#         # Handle cases where there are no family members
#             sheet.cell(row=row_idx, column=25, value="N/A")
#             sheet.cell(row=row_idx, column=26, value="N/A")
#             sheet.cell(row=row_idx, column=27, value="N/A")
#             sheet.cell(row=row_idx, column=28, value="N/A")
#             sheet.cell(row=row_idx, column=29, value="N/A")


#         educational_qualifications = ", ".join([
#             f"{eq.examination_passed} ({eq.year_of_passing})" for eq in profile.educational_qualifications.all()
#         ])
#         sheet.cell(row=row_idx, column=26, value=educational_qualifications)

#         employment_records = ", ".join([
#             f"{er.organization} ({er.designation})" for er in profile.employment_records.all()
#         ])
#         sheet.cell(row=row_idx, column=27, value=employment_records)

#         languages = ", ".join([
#             f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})"
#             for lp in profile.languages.all()
#         ])
#         sheet.cell(row=row_idx, column=28, value=languages)

#         references = ", ".join([
#             f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all()
#         ])
#         sheet.cell(row=row_idx, column=29, value=references)

#         # Compliance details
#         try:
#             compliance = applicant.employee
#             sheet.cell(row=row_idx, column=30, value=compliance.uan_number)
#             sheet.cell(row=row_idx, column=31, value=compliance.pan_number)
#             sheet.cell(row=row_idx, column=32, value=compliance.esic_number)
#         except AttributeError:
#             sheet.cell(row=row_idx, column=30, value="N/A")
#             sheet.cell(row=row_idx, column=31, value="N/A")
#             sheet.cell(row=row_idx, column=32, value="N/A")

#         # Compensation details
#         try:
#             compensation = EmployeeCTC.objects.get(applicant=applicant)
#             sheet.cell(row=row_idx, column=33, value=compensation.salary.gross_salary())
#             sheet.cell(row=row_idx, column=34, value=compensation.contribution.emp_contribution())
#             sheet.cell(row=row_idx, column=35, value=compensation.deduction.emp_deduction())
#             sheet.cell(row=row_idx, column=36, value=compensation.calculate_ctc())
#         except EmployeeCTC.DoesNotExist:
#             sheet.cell(row=row_idx, column=33, value="N/A")
#             sheet.cell(row=row_idx, column=34, value="N/A")
#             sheet.cell(row=row_idx, column=35, value="N/A")
#             sheet.cell(row=row_idx, column=36, value="N/A")

#     # Set the HTTP response with the Excel file
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response



############################### Organize version code ################# 



# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference

# def create_headers(sheet):
#     """Define and set headers for the Excel sheet."""
#     # Merge cells and set section titles
#     sections = [
#         ("Personal Information", 1, 17),
#         ("Banking Information", 18, 21),
#         ("Emergency Contact Information", 22, 24),
#         ("Family Members", 25, 28),
#         ("Educational Qualifications", 29, 29),
#         ("Employment Records", 30, 30),
#         ("Language Proficiencies", 31, 31),
#         ("References", 32, 32),
#         ("Compliance Information", 33, 35),
#         ("Compensation Details", 36, 39),
#     ]

#     for title, start_col, end_col in sections:
#         sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
#         cell = sheet.cell(row=1, column=start_col, value=title)
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Define and write column headers
#     headers = [
#         # Personal Information
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date",
#         "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email",
#         "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address",
#         "Marital Status",

#         # Banking Information
#         "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",

#         # Emergency Contact Information
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",

#         # Family Members
#         "Relation", "Name", "DOB", "Sex",

#         # Educational Qualifications, Employment Records, etc.
#         "Educational Qualifications", "Employment Records", "Language Proficiencies",

#         # References
#         "References",

#         # Compliance Information
#         "UAN Number", "PAN Number (Compliance)", "ESIC Number",

#         # Compensation Details
#         "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)

# def write_applicant_data(sheet, row_idx, applicant):
#     """Write applicant data into the Excel sheet."""
#     profile = applicant.profile

#     # Write personal information
#     personal_data = [
#         profile.fullname, profile.employee_code, profile.designation, profile.date_of_joining, profile.end_date,
#         profile.get_gender_display(), profile.date_of_birth, profile.age, profile.email, profile.phone_number,
#         profile.personal_email, profile.father_name, profile.mother_name, profile.blood_group, profile.present_address,
#         profile.permanent_address, profile.marital_status
#     ]
#     for col_idx, value in enumerate(personal_data, start=1):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write banking information
#     banking_data = [profile.pan_number, profile.bank_name, profile.ifsc_code, profile.bank_account_number]
#     for col_idx, value in enumerate(banking_data, start=18):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write emergency contact information
#     emergency_data = [profile.emergency_contact_name, profile.emergency_contact_relation, profile.emergency_contact_number]
#     for col_idx, value in enumerate(emergency_data, start=22):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write family member information
#     family_member = profile.family_members.first()
#     if family_member:
#         family_data = [family_member.relation, family_member.name, family_member.date_of_birth, family_member.sex]
#     else:
#         family_data = ["N/A"] * 4
#     for col_idx, value in enumerate(family_data, start=25):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # # Write educational qualifications
#     # educational_qualifications = ", ".join([
#     #     f"{eq.examination_passed} ({eq.year_of_passing})"
#     #     for eq in profile.educational_qualifications.all()
#     # ])
#     # sheet.cell(row=row_idx, column=29, value=educational_qualifications)

#         # Write educational qualifications
#     educational_qualifications = profile.educational_qualifications.all()
#     if educational_qualifications.exists():
#         for eq_idx, eq in enumerate(educational_qualifications, start=row_idx):
#             eq_data = [
#                 eq.examination_passed,
#                 eq.year_of_passing,
#                 eq.school_or_college,
#                 eq.subjects,
#                 eq.division
#             ]
#             for col_idx, value in enumerate(eq_data, start=29):
#                 sheet.cell(row=eq_idx, column=col_idx, value=value)
#             row_idx += 1  # Increment the row for each qualification
#     else:
#         sheet.cell(row=row_idx, column=29, value="N/A")

#     # Write employment records
#     employment_records = ", ".join([
#         f"{er.organization} ({er.designation})"
#         for er in profile.employment_records.all()
#     ])
#     sheet.cell(row=row_idx, column=30, value=employment_records)

#     # Write language proficiencies
#     languages = ", ".join([
#         f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})"
#         for lp in profile.languages.all()
#     ])
#     sheet.cell(row=row_idx, column=31, value=languages)

#     # Write references
#     references = ", ".join([
#         f"{ref.name} ({ref.occupation}, {ref.phone_number})"
#         for ref in profile.references.all()
#     ])
#     sheet.cell(row=row_idx, column=32, value=references)

#     # Write compliance information
#     try:
#         compliance = applicant.employee
#         compliance_data = [compliance.uan_number, compliance.pan_number, compliance.esic_number]
#     except AttributeError:
#         compliance_data = ["N/A"] * 3
#     for col_idx, value in enumerate(compliance_data, start=33):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write compensation details
#     # try:
#     #     compensation = applicant.employee.ctc
#     #     compensation_data = [
#     #         compensation.salary.gross_salary(),
#     #         compensation.contribution.emp_contribution(),
#     #         compensation.deduction.emp_deduction(),
#     #         compensation.calculate_ctc()
#     #     ]
#     # except AttributeError:
#     #     compensation_data = ["N/A"] * 4
#     # for col_idx, value in enumerate(compensation_data, start=36):
#     #     sheet.cell(row=row_idx, column=col_idx, value=value)


#         # Write compensation details
#     try:
#         compensation = EmployeeCTC.objects.get(applicant=applicant)
#         compensation_data = [
#             compensation.salary.gross_salary(),
#             compensation.contribution.emp_contribution(),
#             compensation.deduction.emp_deduction(),
#             compensation.calculate_ctc()
#         ]
#     except EmployeeCTC.DoesNotExist:
#         compensation_data = ["N/A"] * 4
#     for col_idx, value in enumerate(compensation_data, start=36):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

# def export_applicants_to_excel(request):
#     """Export applicants' data to an Excel file."""
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     create_headers(sheet)

#     # Fetch applicants and related data
#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members',
#         'profile__educational_qualifications',
#         'profile__employment_records',
#         'profile__languages',
#         'profile__references',
#     )

#     # Write applicant data row by row
#     for row_idx, applicant in enumerate(applicants, start=3):
#         write_applicant_data(sheet, row_idx, applicant)

#     # Set the HTTP response with the Excel file
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response



############################ Organized code based on Education Qualification ##########






# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference

# def create_headers(sheet):
#     """Define and set headers for the Excel sheet."""
#     # Merge cells and set section titles
#     sections = [
#         ("Personal Information", 1, 17),
#         ("Banking Information", 18, 21),
#         ("Emergency Contact Information", 22, 24),
#         ("Family Members", 25, 28),
#         ("Educational Qualifications", 29, 29),
#         ("Employment Records", 30, 30),
#         ("Language Proficiencies", 31, 31),
#         ("References", 32, 32),
#         ("Compliance Information", 33, 35),
#         ("Compensation Details", 36, 39),
#     ]

#     for title, start_col, end_col in sections:
#         sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
#         cell = sheet.cell(row=1, column=start_col, value=title)
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Define and write column headers
#     headers = [
#         # Personal Information
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date",
#         "Gender", "Date of Birth", "Age", "Email", "Phone Number", "Personal Email",
#         "Father Name", "Mother Name", "Blood Group", "Present Address", "Permanent Address",
#         "Marital Status",

#         # Banking Information
#         "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",

#         # Emergency Contact Information
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",

#         # Family Members
#         "Relation", "Name", "DOB", "Sex",

#         # Educational Qualifications, Employment Records, etc.
#         "Educational Qualifications", "Employment Records", "Language Proficiencies",

#         # References
#         "References",

#         # Compliance Information
#         "UAN Number", "PAN Number (Compliance)", "ESIC Number",

#         # Compensation Details
#         "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)

# def write_applicant_data(sheet, row_idx, applicant):
#     """Write applicant data into the Excel sheet."""
#     profile = applicant.profile

#     # Write personal information
#     personal_data = [
#         profile.fullname, profile.employee_code, profile.designation, profile.date_of_joining, profile.end_date,
#         profile.get_gender_display(), profile.date_of_birth, profile.age, profile.email, profile.phone_number,
#         profile.personal_email, profile.father_name, profile.mother_name, profile.blood_group, profile.present_address,
#         profile.permanent_address, profile.marital_status
#     ]
#     for col_idx, value in enumerate(personal_data, start=1):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write banking information
#     banking_data = [profile.pan_number, profile.bank_name, profile.ifsc_code, profile.bank_account_number]
#     for col_idx, value in enumerate(banking_data, start=18):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write emergency contact information
#     emergency_data = [profile.emergency_contact_name, profile.emergency_contact_relation, profile.emergency_contact_number]
#     for col_idx, value in enumerate(emergency_data, start=22):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write family member information
#     family_member = profile.family_members.first()
#     if family_member:
#         family_data = [family_member.relation, family_member.name, family_member.date_of_birth, family_member.sex]
#     else:
#         family_data = ["N/A"] * 4
#     for col_idx, value in enumerate(family_data, start=25):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # # Write educational qualifications
#     # educational_qualifications = ", ".join([
#     #     f"{eq.examination_passed} ({eq.year_of_passing})"
#     #     for eq in profile.educational_qualifications.all()
#     # ])
#     # sheet.cell(row=row_idx, column=29, value=educational_qualifications)

#         # Write educational qualifications
#     educational_qualifications = profile.educational_qualifications.all()
#     if educational_qualifications.exists():
#         for eq_idx, eq in enumerate(educational_qualifications, start=row_idx):
#             eq_data = [
#                 eq.examination_passed,
#                 eq.year_of_passing,
#                 eq.school_or_college,
#                 eq.subjects,
#                 eq.division
#             ]
#             for col_idx, value in enumerate(eq_data, start=29):
#                 sheet.cell(row=eq_idx, column=col_idx, value=value)
#             row_idx += 1  # Increment the row for each qualification
#     else:
#         sheet.cell(row=row_idx, column=29, value="N/A")

#     # Write employment records
#     employment_records = ", ".join([
#         f"{er.organization} ({er.designation})"
#         for er in profile.employment_records.all()
#     ])
#     sheet.cell(row=row_idx, column=30, value=employment_records)

#     # Write language proficiencies
#     languages = ", ".join([
#         f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})"
#         for lp in profile.languages.all()
#     ])
#     sheet.cell(row=row_idx, column=31, value=languages)

#     # Write references
#     references = ", ".join([
#         f"{ref.name} ({ref.occupation}, {ref.phone_number})"
#         for ref in profile.references.all()
#     ])
#     sheet.cell(row=row_idx, column=32, value=references)

#     # Write compliance information
#     try:
#         compliance = applicant.employee
#         compliance_data = [compliance.uan_number, compliance.pan_number, compliance.esic_number]
#     except AttributeError:
#         compliance_data = ["N/A"] * 3
#     for col_idx, value in enumerate(compliance_data, start=33):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Write compensation details
#     # try:
#     #     compensation = applicant.employee.ctc
#     #     compensation_data = [
#     #         compensation.salary.gross_salary(),
#     #         compensation.contribution.emp_contribution(),
#     #         compensation.deduction.emp_deduction(),
#     #         compensation.calculate_ctc()
#     #     ]
#     # except AttributeError:
#     #     compensation_data = ["N/A"] * 4
#     # for col_idx, value in enumerate(compensation_data, start=36):
#     #     sheet.cell(row=row_idx, column=col_idx, value=value)


#         # Write compensation details
#     try:
#         compensation = EmployeeCTC.objects.get(applicant=applicant)
#         compensation_data = [
#             compensation.salary.gross_salary(),
#             compensation.contribution.emp_contribution(),
#             compensation.deduction.emp_deduction(),
#             compensation.calculate_ctc()
#         ]
#     except EmployeeCTC.DoesNotExist:
#         compensation_data = ["N/A"] * 4
#     for col_idx, value in enumerate(compensation_data, start=36):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

# def export_applicants_to_excel(request):
#     """Export applicants' data to an Excel file."""
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     create_headers(sheet)

#     # Fetch applicants and related data
#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members',
#         'profile__educational_qualifications',
#         'profile__employment_records',
#         'profile__languages',
#         'profile__references',
#     )

#     # Write applicant data row by row
#     for row_idx, applicant in enumerate(applicants, start=3):
#         write_applicant_data(sheet, row_idx, applicant)

#     # Set the HTTP response with the Excel file
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response





####################### EDsn Qlfcn V2 ###################




# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, EmployeeCTC

# def create_headers(sheet):
#     """Set headers for the Excel sheet."""
#     sections = [
#         ("Personal Information", 1, 17), ("Banking Information", 18, 21),
#         ("Emergency Contact Information", 22, 24), ("Family Members", 25, 28),
#         ("Educational Qualifications", 29, 33), ("Employment Records", 34, 34),
#         ("Language Proficiencies", 35, 35), ("References", 36, 36),
#         ("Compliance Information", 37, 39), ("Compensation Details", 40, 43),
#     ]
#     for title, start_col, end_col in sections:
#         sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
#         sheet.cell(row=1, column=start_col, value=title).alignment = Alignment(horizontal="center")
#     headers = [
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date", "Gender", "Date of Birth",
#         "Age", "Email", "Phone Number", "Personal Email", "Father Name", "Mother Name", "Blood Group", "Present Address",
#         "Permanent Address", "Marital Status", "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number", "Relation", "Name", "DOB",
#         "Sex", "Examination Passed", "Year of Passing", "School/College", "Subjects", "Division", "Employment Records",
#         "Language Proficiencies", "References", "UAN Number", "PAN Number (Compliance)", "ESIC Number", "Gross Salary",
#         "Total Contribution", "Total Deductions", "Net CTC"
#     ]
#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)

# def write_applicant_data(sheet, row_idx, applicant):
#     """Write applicant data into the Excel sheet."""
#     profile = applicant.profile
#     personal_data = [
#         profile.fullname, profile.employee_code, profile.designation, profile.date_of_joining, profile.end_date,
#         profile.get_gender_display(), profile.date_of_birth, profile.age, profile.email, profile.phone_number,
#         profile.personal_email, profile.father_name, profile.mother_name, profile.blood_group, profile.present_address,
#         profile.permanent_address, profile.marital_status
#     ]
#     banking_data = [profile.pan_number, profile.bank_name, profile.ifsc_code, profile.bank_account_number]
#     emergency_data = [profile.emergency_contact_name, profile.emergency_contact_relation, profile.emergency_contact_number]
#     family_member = profile.family_members.first()
#     family_data = [family_member.relation, family_member.name, family_member.date_of_birth, family_member.sex] if family_member else ["N/A"] * 4
#     qualifications = profile.educational_qualifications.all()
#     education_data = [[q.examination_passed, q.year_of_passing, q.school_or_college, q.subjects, q.division] for q in qualifications] if qualifications else [["N/A"] * 5]
#     employment_records = ", ".join(f"{er.organization} ({er.designation})" for er in profile.employment_records.all())
#     languages = ", ".join(f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})" for lp in profile.languages.all())
#     references = ", ".join(f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all())
#     compliance = getattr(applicant.employee, 'uan_number', "N/A"), getattr(applicant.employee, 'pan_number', "N/A"), getattr(applicant.employee, 'esic_number', "N/A")
#     try:
#         compensation = EmployeeCTC.objects.get(applicant=applicant)
#         compensation_data = [
#             compensation.salary.gross_salary(), compensation.contribution.emp_contribution(),
#             compensation.deduction.emp_deduction(), compensation.calculate_ctc()
#         ]
#     except EmployeeCTC.DoesNotExist:
#         compensation_data = ["N/A"] * 4

#     all_data = [personal_data, banking_data, emergency_data, family_data, [employment_records], [languages], [references], compliance, compensation_data]
#     col = 1
#     for data_set in all_data:
#         for value in data_set:
#             if isinstance(value, list):
#                 for sub_value in value:
#                     sheet.cell(row=row_idx, column=col, value=sub_value)
#                     col += 1
#             else:
#                 sheet.cell(row=row_idx, column=col, value=value)
#                 col += 1

#     for q_row in education_data:
#         for idx, val in enumerate(q_row, start=29):
#             sheet.cell(row=row_idx, column=idx, value=val)
#         row_idx += 1  # Increment the row for additional qualifications

# def export_applicants_to_excel(request):
#     """Export applicants' data to an Excel file."""
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"
#     create_headers(sheet)
#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members', 'profile__educational_qualifications', 'profile__employment_records',
#         'profile__languages', 'profile__references',
#     )
#     row_idx = 3
#     for applicant in applicants:
#         write_applicant_data(sheet, row_idx, applicant)
#         row_idx += 1
#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response


############################## 



# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference 
# from compensation.models import EmployeeCTC

# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# # from users.models import Profile, EmployeeCTC

# # Create headers for the Excel sheet
# def create_headers(sheet):
#     sections = [
#         ("Personal Information", 1, 17),
#         ("Banking Information", 18, 21),
#         ("Emergency Contact Information", 22, 24),
#         ("Family Members", 25, 28),
#         ("Educational Qualifications", 29, 33),
#         ("Employment Records", 34, 34),
#         ("Language Proficiencies", 35, 35),
#         ("References", 36, 36),
#         ("Compliance Information", 37, 39),
#         ("Compensation Details", 40, 43),
#     ]

#     for title, start_col, end_col in sections:
#         sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
#         cell = sheet.cell(row=1, column=start_col, value=title)
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     headers = [
#         # Personal Information
#         "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date", "Gender", "Date of Birth", 
#         "Age", "Email", "Phone Number", "Personal Email", "Father Name", "Mother Name", "Blood Group", "Present Address", 
#         "Permanent Address", "Marital Status",
#         # Banking Information
#         "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",
#         # Emergency Contact Information
#         "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",
#         # Family Members
#         "Relation", "Name", "DOB", "Sex",
#         # Educational Qualifications
#         "Examination Passed", "Year of Passing", "School/College", "Subjects", "Division",
#         # Employment Records
#         "Employment Records",
#         # Language Proficiencies
#         "Language Proficiencies",
#         # References
#         "References",
#         # Compliance Information
#         "UAN Number", "PAN Number (Compliance)", "ESIC Number",
#         # Compensation Details
#         "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)

# # Write applicant data to the Excel sheet
# def write_applicant_data(sheet, row_idx, applicant):
#     profile = applicant.profile

#     # Personal Information
#     personal_data = [
#         profile.fullname, profile.employee_code, profile.designation, profile.date_of_joining, profile.end_date,
#         profile.get_gender_display(), profile.date_of_birth, profile.age, profile.email, profile.phone_number,
#         profile.personal_email, profile.father_name, profile.mother_name, profile.blood_group, profile.present_address,
#         profile.permanent_address, profile.marital_status
#     ]
#     for col_idx, value in enumerate(personal_data, start=1):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Banking Information
#     banking_data = [profile.pan_number, profile.bank_name, profile.ifsc_code, profile.bank_account_number]
#     for col_idx, value in enumerate(banking_data, start=18):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Emergency Contact Information
#     emergency_data = [profile.emergency_contact_name, profile.emergency_contact_relation, profile.emergency_contact_number]
#     for col_idx, value in enumerate(emergency_data, start=22):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Family Members
#     family_member = profile.family_members.first()
#     family_data = [family_member.relation, family_member.name, family_member.date_of_birth, family_member.sex] if family_member else ["N/A"] * 4
#     for col_idx, value in enumerate(family_data, start=25):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Educational Qualifications
#     qualifications = profile.educational_qualifications.all()
#     if qualifications.exists():
#         for q in qualifications:
#             q_data = [q.examination_passed, q.year_of_passing, q.school_or_college, q.subjects, q.division]
#             for col_idx, value in enumerate(q_data, start=29):
#                 sheet.cell(row=row_idx, column=col_idx, value=value)
#             row_idx += 1
#     else:
#         for col_idx in range(29, 34):
#             sheet.cell(row=row_idx, column=col_idx, value="N/A")

#     # Employment Records
#     employment_records = ", ".join(f"{er.organization} ({er.designation})" for er in profile.employment_records.all())
#     sheet.cell(row=row_idx, column=34, value=employment_records)

#     # Language Proficiencies
#     languages = ", ".join(f"{lp.language} (Speak: {lp.speak}, Read: {lp.read}, Write: {lp.write})" for lp in profile.languages.all())
#     sheet.cell(row=row_idx, column=35, value=languages)

#     # References
#     references = ", ".join(f"{ref.name} ({ref.occupation}, {ref.phone_number})" for ref in profile.references.all())
#     sheet.cell(row=row_idx, column=36, value=references)

#     # Compliance Information
#     compliance_data = [getattr(applicant.employee, attr, "N/A") for attr in ["uan_number", "pan_number", "esic_number"]]
#     for col_idx, value in enumerate(compliance_data, start=37):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Compensation Details
#     try:
#         compensation = EmployeeCTC.objects.get(applicant=applicant)
#         compensation_data = [
#             compensation.salary.gross_salary(), compensation.contribution.emp_contribution(),
#             compensation.deduction.emp_deduction(), compensation.calculate_ctc()
#         ]
#     except EmployeeCTC.DoesNotExist:
#         compensation_data = ["N/A"] * 4
#     for col_idx, value in enumerate(compensation_data, start=40):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

# # # Export applicants to Excel
# # def export_applicants_to_excel(request):
# #     workbook = Workbook()
# #     sheet = workbook.active
# #     sheet.title = "Applicants Data"

# #     create_headers(sheet)

# #     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
# #         'profile__family_members', 'profile__educational_qualifications', 'profile__employment_records',
# #         'profile__languages', 'profile__references',
# #     )

# #     row_idx = 3
# #     for applicant in applicants:
# #         write_applicant_data(sheet, row_idx, applicant)
# #         row_idx += 1

# #     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
# #     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
# #     workbook.save(response)
# #     return response




# def export_applicants_to_excel(request):
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     create_headers(sheet)

#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members', 'profile__educational_qualifications', 'profile__employment_records',
#         'profile__languages', 'profile__references',
#     )

#     row_idx = 3
#     for applicant in applicants:
#         try:
#             write_applicant_data(sheet, row_idx, applicant)
#         except Exception as e:
#             print(f"Error processing applicant {applicant}: {e}")
#         row_idx += 1

#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response





######################### Organized code version   ###############



from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from applicant.models import Applicant
from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference
from compensation.models import EmployeeCTC


# Create headers for the Excel sheet
def create_headers(sheet):
    sections = [
        ("Personal Information", 1, 17),
        ("Banking Information", 18, 21),
        ("Emergency Contact Information", 22, 24),
        ("Family Members", 25, 28),
        ("Educational Qualifications", 29, 33),
        ("Employment Records", 34, 38),  # Expanded columns for subheaders
        ("Language Proficiencies", 39, 42),  # Expanded columns for subheaders
        ("References", 43, 46),  # Expanded columns for subheaders
        ("Compliance Information", 47, 49),
        ("Compensation Details", 50, 53),
    ]

    # Merge cells for section titles
    for title, start_col, end_col in sections:
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = sheet.cell(row=1, column=start_col, value=title)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add column headers
    headers = [
        # Personal Information
        "Applicant Name", "Employee Code", "Designation", "Date of Joining", "End Date", "Gender", "Date of Birth",
        "Age", "Email", "Phone Number", "Personal Email", "Father Name", "Mother Name", "Blood Group",
        "Present Address", "Permanent Address", "Marital Status",
        # Banking Information
        "PAN Number", "Bank Name", "IFSC Code", "Bank Account Number",
        # Emergency Contact Information
        "Emergency Contact Name", "Emergency Contact Relation", "Emergency Contact Number",
        # Family Members
        "Relation", "Name", "DOB", "Sex",
        # Educational Qualifications
        "Examination Passed", "Year of Passing", "School/College", "Subjects", "Division",
        # Employment Records (Subheaders)
        "Organization", "Designation", "Joining Date", "Leaving Date", "Document",
        # Language Proficiencies (Subheaders)
        "Language", "Speak", "Read", "Write",
        # References (Subheaders)
        "Name", "Occupation", "Phone Number", "Email",
        # Compliance Information
        "UAN Number", "PAN Number (Compliance)", "ESIC Number",
        # Compensation Details
        "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
    ]

    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)



def write_section_data(sheet, start_row, items, start_col, data_extractor):
    """
    Writes section data into the Excel sheet.
    - `items`: QuerySet or list of objects to write.
    - `data_extractor`: Function to extract the row's data.
    """
    if items.exists():
        for item in items:
            data = data_extractor(item)
            for col_idx, value in enumerate(data, start=start_col):
                sheet.cell(row=start_row, column=col_idx, value=value)
            start_row += 1  # Increment only when data is written
    return start_row  # Return the row index (unchanged if no data was written)




def write_applicant_data(sheet, start_row, applicant):
    """
    Writes applicant data to the Excel sheet, including expandable sections.
    """
    profile = getattr(applicant, "profile", None)
    current_row = start_row

    # Default empty values for fields in case `profile` does not exist
    if not profile:
        profile_data = ["N/A"] * 17
    else:
        profile_data = [
            profile.fullname, profile.employee_code, profile.designation, profile.date_of_joining, profile.end_date,
            profile.get_gender_display(), profile.date_of_birth, profile.age, profile.email, profile.phone_number,
            profile.personal_email, profile.father_name, profile.mother_name, profile.blood_group, profile.present_address,
            profile.permanent_address, profile.marital_status
        ]


    # Write Personal Information
    for col_idx, value in enumerate(profile_data, start=1):
        sheet.cell(row=current_row, column=col_idx, value=value)

    # Fetch related data for expandable sections
    qualifications = list(profile.educational_qualifications.all()) if profile else []
    employment_records = list(profile.employment_records.all()) if profile else []
    languages = list(profile.languages.all()) if profile else []
    references = list(profile.references.all()) if profile else []

    # Banking Information
    banking_data = [
        profile.pan_number if profile else "N/A",
        profile.bank_name if profile else "N/A",
        profile.ifsc_code if profile else "N/A",
        profile.bank_account_number if profile else "N/A",
    ]
    for col_idx, value in enumerate(banking_data, start=18):
        sheet.cell(row=current_row, column=col_idx, value=value)

    # Emergency Contact Information
    emergency_data = [
        profile.emergency_contact_name if profile else "N/A",
        profile.emergency_contact_relation if profile else "N/A",
        profile.emergency_contact_number if profile else "N/A",
    ]
    for col_idx, value in enumerate(emergency_data, start=22):
        sheet.cell(row=current_row, column=col_idx, value=value)

    # Family Members
    family_members = profile.family_members.all() if profile else []
    if family_members:
        for fm in family_members:
            family_data = [fm.relation, fm.name, fm.date_of_birth, fm.sex]
            for col_idx, value in enumerate(family_data, start=25):
                sheet.cell(row=current_row, column=col_idx, value=value)
            current_row += 1
    else:
        family_data = ["N/A"] * 4
        for col_idx, value in enumerate(family_data, start=25):
            sheet.cell(row=current_row, column=col_idx, value=value)
        current_row += 1


    # Fetch compliance and compensation information
    compliance_data = ["N/A"] * 3
    if hasattr(applicant, "employee"):
        compliance_data = [
            getattr(applicant.employee, attr, "N/A")
            for attr in ["uan_number", "pan_number", "esic_number"]
        ]

    compensation_data = ["N/A"] * 4
    try:
        compensation = EmployeeCTC.objects.get(applicant=applicant)
        compensation_data = [
            compensation.salary.gross_salary(),
            compensation.contribution.emp_contribution(),
            compensation.deduction.emp_deduction(),
            compensation.calculate_ctc(),
        ]
    except EmployeeCTC.DoesNotExist:
        pass

    # Determine maximum rows required for expandable sections
    max_rows = max(len(qualifications), len(employment_records), len(languages), len(references), 1)

    # Write expandable sections
    for i in range(max_rows):
        row_data = []

        # Educational Qualifications
        if i < len(qualifications):
            q = qualifications[i]
            row_data.extend([q.examination_passed, q.year_of_passing, q.school_or_college, q.subjects, q.division])
        else:
            row_data.extend([""] * 5)

        # Employment Records
        if i < len(employment_records):
            er = employment_records[i]
            row_data.extend([er.organization, er.designation, er.joining_date, er.leaving_date, er.document or "N/A"])
        else:
            row_data.extend([""] * 5)

        # Language Proficiencies
        if i < len(languages):
            lp = languages[i]
            row_data.extend([lp.language, lp.speak, lp.read, lp.write])
        else:
            row_data.extend([""] * 4)

        # References
        if i < len(references):
            ref = references[i]
            row_data.extend([ref.name, ref.occupation, ref.phone_number, ref.email])
        else:
            row_data.extend([""] * 4)

        # Compliance Information (only for the first row)
        if i == 0:
            row_data.extend(compliance_data)
        else:
            row_data.extend([""] * 3)

        # Compensation Details (only for the first row)
        if i == 0:
            row_data.extend(compensation_data)
        else:
            row_data.extend([""] * 4)

        # Write the row data
        for col_idx, value in enumerate(row_data, start=29):  # Columns for expandable sections
            sheet.cell(row=current_row, column=col_idx, value=value)

        current_row += 1

    return current_row  # Return the next available row






 
# Export applicants to Excel
def export_applicants_to_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applicants Data"

    # Create headers
    create_headers(sheet)

    # Fetch all applicants and related data
    applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
        'profile__family_members', 'profile__educational_qualifications',
        'profile__employment_records', 'profile__languages', 'profile__references',
    )

    row_idx = 3  # Start writing below the headers
    for applicant in applicants:
        try:
            row_idx = write_applicant_data(sheet, row_idx, applicant)
        except Exception as e:
            print(f"Error processing applicant {applicant.id}: {e}")

    # Save and return response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
    workbook.save(response)
    return response
