from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth import logout
from .forms import CustomAuthenticationForm 
from compliance.models import Employee
from compensation.models import EmployeeCTC
import openpyxl
from django.http import HttpResponse
from users.models import Profile
from applicant.models import Applicant
from django.utils.dateparse import parse_date
from django.db.models import Q
from io import BytesIO
import openpyxl
from .models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request=request, data=request.POST)
        if form.is_valid():
            login(request, form.user_cache)
            try:
                if request.user.applicant:  # Trying to access the related applicant
                    return redirect('applicant:dashboard')
            except request.user._meta.model.applicant.RelatedObjectDoesNotExist:
                return redirect('core:main_page')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'users/login.html', {'form': form})

def logout_view(request):
    logout(request)

    return redirect('users:login_view')

########################### Modifying code Personal Information for columns wise

def export_profile_to_excel(profile, family_members, educational_qualifications, employment_records, language_proficiencies, references, compliance_info, compensation_info):
    import openpyxl
    from openpyxl.styles import Alignment
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Profile Details"

    # Define headers for the personal details section
    personal_headers = [
        "Employee Code", "First Name", "Middle Name", "Last Name",
        "Designation", "Date of Joining", "End Date", "Email",
        "Phone Number", "Gender", "Date of Birth", "Present Address",
        "Permanent Address", "Personal Email", "PAN Number", "Marital Status"
    ]
    personal_details = [
        profile.employee_code, profile.first_name, profile.middle_name,
        profile.last_name, profile.designation, profile.date_of_joining,
        profile.end_date, profile.email, profile.phone_number,
        profile.get_gender_display(), profile.date_of_birth,
        profile.present_address, profile.permanent_address,
        profile.personal_email, profile.pan_number, profile.marital_status
    ]
    # Write headers and personal details in column-wise format
    for col, header in enumerate(personal_headers, start=1):
        sheet.cell(row=1, column=col, value=header)  # Header row
        sheet.cell(row=2, column=col, value=personal_details[col - 1])  # Data row

    # Adjust alignment for readability
    for col in range(1, len(personal_headers) + 1):
        sheet.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        sheet.cell(row=2, column=col).alignment = Alignment(horizontal='center')

    # Add additional sections for family members, education, etc.
    row_offset = 4

    # Family Members Section
    sheet.cell(row=row_offset, column=1, value="Family Members")
    row_offset += 1
    sheet.append(["Relation", "Name", "Date of Birth", "Sex", "Age"])
    for member in family_members:
        sheet.append([member.relation, member.name, member.date_of_birth, member.sex, member.age])
    row_offset += len(family_members) + 2

    # Educational Qualifications
    sheet.cell(row=row_offset, column=1, value="Educational Qualifications")
    row_offset += 1
    sheet.append(["Examination Passed", "Year of Passing", "School/College", "Division"])
    for qualification in educational_qualifications:
        sheet.append([
            qualification.examination_passed,
            qualification.year_of_passing,
            qualification.school_or_college,
            qualification.division,
        ])
    row_offset += len(educational_qualifications) + 2

    # Employment Records
    sheet.cell(row=row_offset, column=1, value="Employment Records")
    row_offset += 1
    sheet.append(["Organization", "Designation", "Joining Date", "Leaving Date"])
    for record in employment_records:
        sheet.append([
            record.organization, record.designation,
            record.joining_date, record.leaving_date
        ])
    row_offset += len(employment_records) + 2

    # Language Proficiencies
    sheet.cell(row=row_offset, column=1, value="Languages Known")
    row_offset += 1
    sheet.append(["Language", "Speak", "Read", "Write"])
    for language in language_proficiencies:
        sheet.append([
            language.language,
            language.get_speak_display(),
            language.get_read_display(),
            language.get_write_display()
        ])
    row_offset += len(language_proficiencies) + 2

    # References
    sheet.cell(row=row_offset, column=1, value="References")
    row_offset += 1
    sheet.append(["Name", "Occupation", "Phone Number", "Email"])
    for reference in references:
        sheet.append([
            reference.name, reference.occupation,
            reference.phone_number, reference.email
        ])
    row_offset += len(references) + 2

    # Compliance Information
    sheet.cell(row=row_offset, column=1, value="Compliance Information")
    row_offset += 1
    if compliance_info:
        sheet.append(["UAN Number", compliance_info.uan_number])
        sheet.append(["PAN Number", compliance_info.pan_number])
        sheet.append(["ESIC Number", compliance_info.esic_number])
    else:
        sheet.append(["No compliance data available"])
    row_offset += 3

    # Compensation Information
    sheet.cell(row=row_offset, column=1, value="Compensation Information")
    row_offset += 1
    if compensation_info:
        sheet.append(["Component", "Amount"])
        salary = compensation_info.salary
        contribution = compensation_info.contribution
        deduction = compensation_info.deduction
        sheet.append(["Basic Salary", salary.basic])
        sheet.append(["HRA", salary.hra])
        sheet.append(["Special Allowance", salary.special_allowance])
        sheet.append(["Gross Salary", salary.gross_salary()])
        sheet.append(["Total Contributions", contribution.emp_contribution()])
        sheet.append(["Total Deductions", deduction.emp_deduction()])
        sheet.append(["Net Salary (CTC)", compensation_info.calculate_ctc()])
    else:
        sheet.append(["No compensation data available"])

    return workbook

################################ update views for compliance data
def profile_detail_view(request, profile_id):
    # Get the specific profile by its ID
    profile = get_object_or_404(Profile, id=profile_id)
    all_profiles = Profile.objects.all()

    # Access related FamilyMember, EducationalQualification, etc.
    family_members = profile.family_members.all()
    educational_qualifications = profile.educational_qualifications.all()
    employment_records = profile.employment_records.all()
    language_proficiencies = profile.languages.all()
    references = profile.references.all()

    # Fetch the compliance data through the Applicant linked to this Profile
    applicant = getattr(profile.user, 'applicant', None)  # Access Applicant via the CustomUser relationship
    compliance_info = Employee.objects.filter(applicant=applicant).first() if applicant else None
    compensation_info = EmployeeCTC.objects.filter(applicant=applicant).first()
    # compensation_info = EmployeeCTC.objects.filter(profile=profile).first()

    # Export to Excel if requested
    if request.GET.get('export') == 'excel':
        workbook = export_profile_to_excel(
            profile,
            family_members,
            educational_qualifications,
            employment_records,
            language_proficiencies,
            references,
            compliance_info,
            compensation_info  # Add these two arguments
        )
    # Prepare the response
        response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Profile_{profile.employee_code}.xlsx'
        workbook.save(response)
        return response
    # Pass the data to the template if not exporting
    context = {
        'profile': profile,
        'all_profiles': all_profiles,
        'family_members': family_members,
        'educational_qualifications': educational_qualifications,
        'employment_records': employment_records,
        'language_proficiencies': language_proficiencies,
        'references': references,
        'compliance_info': compliance_info,
        'compensation_info': compensation_info,
    }

    return render(request, 'profile_detail.html', context)

def hiring_report(request):
    from_date = request.GET.get('from_date')
    date_of_joining = request.GET.get('date_of_joining')
    
    profiles = Profile.objects.all()
    
    # Apply the filter based on 'date_of_joining' range
    if from_date and date_of_joining:
        profiles = profiles.filter(
            date_of_joining__range=(parse_date(from_date), parse_date(date_of_joining))
        )
    elif from_date:
        profiles = profiles.filter(date_of_joining__gte=parse_date(from_date))
    elif date_of_joining:
        profiles = profiles.filter(date_of_joining__lte=parse_date(date_of_joining))
    
    return render(request, 'users/hiring_report.html', {'profiles': profiles, 'from_date': from_date, 'date_of_joining': date_of_joining})


def attrition_report(request):
    from_date = request.GET.get('from_date')
    end_date = request.GET.get('end_date')
    
    profiles = Profile.objects.filter(end_date__isnull=False)  # Only include profiles with an end_date

    # Apply the filter based only on the 'end_date' range
    if from_date and end_date:
        profiles = profiles.filter(
            end_date__range=(parse_date(from_date), parse_date(end_date))
        )
    elif from_date:
        profiles = profiles.filter(end_date__gte=parse_date(from_date))
    elif end_date:
        profiles = profiles.filter(end_date__lte=parse_date(end_date))
    
    return render(request, 'users/attrition_report.html', {'profiles': profiles, 'from_date': from_date, 'end_date': end_date})


######################### Organized code version   ###############



# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference
# from compensation.models import EmployeeCTC


# # Create headers for the Excel sheet
# def create_headers(sheet):
#     sections = [
#         ("Educational Qualifications", 29, 33),
#         ("Employment Records", 34, 38),  # Expanded columns for subheaders
#         ("Language Proficiencies", 39, 42),  # Expanded columns for subheaders
#         ("References", 43, 46),  # Expanded columns for subheaders
#         ("Compliance Information", 47, 49),
#         ("Compensation Details", 50, 53),
#     ]

#     # Merge cells for section titles
#     for title, start_col, end_col in sections:
#         sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
#         cell = sheet.cell(row=1, column=start_col, value=title)
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Add column headers
#     headers = [
#         # Educational Qualifications
#         "Examination Passed", "Year of Passing", "School/College", "Subjects", "Division",
#         # Employment Records (Subheaders)
#         "Organization", "Designation", "Joining Date", "Leaving Date", "Document",
#         # Language Proficiencies (Subheaders)
#         "Language", "Speak", "Read", "Write",
#         # References (Subheaders)
#         "Name", "Occupation", "Phone Number", "Email",
#         # Compliance Information
#         "UAN Number", "PAN Number (Compliance)", "ESIC Number",
#         # Compensation Details
#         "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)


# def write_applicant_data(sheet, row_idx, applicant):
#     profile = applicant.profile
#     current_row = row_idx  # Start from the given row
#     next_row = current_row  # Track the next row for expandable sections

#     # Educational Qualifications

#     qualifications = profile.educational_qualifications.all()
#     # Start writing data below the header row (e.g., if headers are in row 2, data starts at row 3)
#     start_row = next_row  # Start row below headers
#     # next_row = header_row + 1 
#     if qualifications.exists():
#         for q in qualifications:
#         # Extract qualification data
#             q_data = [q.examination_passed, q.year_of_passing, q.school_or_college, q.subjects, q.division]
        
#         # Write each field in its respective column
#             for col_idx, value in enumerate(q_data, start=29):
#                 sheet.cell(row=start_row, column=col_idx, value=value)
        
#         # Move to the next row for the next qualification
#             start_row += 1
#     else:
#     # Write "N/A" if no qualifications exist
#         q_data = ["N/A"] * 5
#         for col_idx, value in enumerate(q_data, start=29):
#             sheet.cell(row=start_row, column=col_idx, value=value)
#         start_row += 1  # Ensure the next section starts from a new row
#     next_row=start_row

# # Update the next_row tracker for the subsequent sections
#     next_row = start_row    


#     # Employment Records
#     employment_records = profile.employment_records.all()
#     if employment_records.exists():
#         for er in employment_records:
#             er_data = [er.organization, er.designation, er.joining_date, er.leaving_date, er.document or "N/A"]
#             for col_idx, value in enumerate(er_data, start=34):
#                 sheet.cell(row=next_row, column=col_idx, value=value)
#             next_row += 1
#     else:
#         er_data = ["N/A"] * 5
#         for col_idx, value in enumerate(er_data, start=34):
#             sheet.cell(row=next_row, column=col_idx, value=value)
#         next_row += 1

#     # Language Proficiencies
#     languages = profile.languages.all()
#     if languages.exists():
#         for lp in languages:
#             lp_data = [lp.language, lp.speak, lp.read, lp.write]
#             for col_idx, value in enumerate(lp_data, start=39):
#                 sheet.cell(row=next_row, column=col_idx, value=value)
#             next_row += 1
#     else:
#         lp_data = ["N/A"] * 4
#         for col_idx, value in enumerate(lp_data, start=39):
#             sheet.cell(row=next_row, column=col_idx, value=value)
#         next_row += 1

#     # References
#     references = profile.references.all()
#     if references.exists():
#         for ref in references:
#             ref_data = [ref.name, ref.occupation, ref.phone_number, ref.email]
#             for col_idx, value in enumerate(ref_data, start=43):
#                 sheet.cell(row=next_row, column=col_idx, value=value)
#             next_row += 1
#     else:
#         ref_data = ["N/A"] * 4
#         for col_idx, value in enumerate(ref_data, start=43):
#             sheet.cell(row=next_row, column=col_idx, value=value)
#         next_row += 1

#     # Compliance Information
#     compliance_data = [getattr(applicant.employee, attr, "N/A") for attr in ["uan_number", "pan_number", "esic_number"]]
#     for col_idx, value in enumerate(compliance_data, start=47):
#         sheet.cell(row=current_row, column=col_idx, value=value)

#     # Compensation Details
#     try:
#         compensation = EmployeeCTC.objects.get(applicant=applicant)
#         compensation_data = [
#             compensation.salary.gross_salary(), compensation.contribution.emp_contribution(),
#             compensation.deduction.emp_deduction(), compensation.calculate_ctc()
#         ]
#     except EmployeeCTC.DoesNotExist:
#         compensation_data = ["N/A"] * 4
#     for col_idx, value in enumerate(compensation_data, start=50):
#         sheet.cell(row=current_row, column=col_idx, value=value)

#     # Return the next available row index
#     return next_row

# # Export applicants to Excel
# def export_applicants_to_excel(request):
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     create_headers(sheet)

#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__family_members', 'profile__educational_qualifications', 'profile__employment_records',
#         'profile__languages', 'profile__references',
#     )

#     row_idx = 3
#     for applicant in applicants:
#         try:
#             write_applicant_data(sheet, row_idx, applicant)
#         except Exception as e:
#             print(f"Error processing applicant {applicant}: {e}")
#         row_idx += 1

#     response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response


#####################################

# from django.http import HttpResponse
# from openpyxl import Workbook
# from openpyxl.styles import Alignment
# from applicant.models import Applicant
# from users.models import Profile, FamilyMember, EducationalQualification, EmploymentRecord, LanguageProficiency, Reference
# from compensation.models import EmployeeCTC


# # Create headers for the Excel sheet
# def create_headers(sheet):
#     sections = [
#         ("Educational Qualifications", 1, 5),
#         ("Employment Records", 6, 10),
#         ("Language Proficiencies", 11, 14),
#         ("References", 15, 18),
#         ("Compliance Information", 19, 21),
#         ("Compensation Details", 22, 25),
#     ]

#     # Merge cells for section titles
#     for title, start_col, end_col in sections:
#         sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
#         cell = sheet.cell(row=1, column=start_col, value=title)
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     # Add column headers
#     headers = [
#         # Educational Qualifications
#         "Examination Passed", "Year of Passing", "School/College", "Subjects", "Division",
#         # Employment Records
#         "Organization", "Designation", "Joining Date", "Leaving Date", "Document",
#         # Language Proficiencies
#         "Language", "Speak", "Read", "Write",
#         # References
#         "Name", "Occupation", "Phone Number", "Email",
#         # Compliance Information
#         "UAN Number", "PAN Number", "ESIC Number",
#         # Compensation Details
#         "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC"
#     ]

#     for col_idx, header in enumerate(headers, start=1):
#         sheet.cell(row=2, column=col_idx, value=header)


# # Write applicant data to the Excel sheet
# def write_applicant_data(sheet, row_idx, applicant):
#     profile = applicant.profile
#     next_row = row_idx

#     # Educational Qualifications
#     qualifications = profile.educational_qualifications.all()
#     next_row = write_section_data(
#         sheet, next_row, qualifications, 1,
#         lambda q: [q.examination_passed, q.year_of_passing, q.school_or_college, q.subjects, q.division]
#     )

#     # Employment Records
#     employment_records = profile.employment_records.all()
#     next_row = write_section_data(
#         sheet, next_row, employment_records, 6,
#         lambda er: [er.organization, er.designation, er.joining_date, er.leaving_date, er.document or "N/A"]
#     )

#     # Language Proficiencies
#     languages = profile.languages.all()
#     next_row = write_section_data(
#         sheet, next_row, languages, 11,
#         lambda lp: [lp.language, lp.speak, lp.read, lp.write]
#     )

#     # References
#     references = profile.references.all()
#     next_row = write_section_data(
#         sheet, next_row, references, 15,
#         lambda ref: [ref.name, ref.occupation, ref.phone_number, ref.email]
#     )

#     # Compliance Information
#     compliance_data = [
#         getattr(applicant.employee, attr, "N/A")
#         for attr in ["uan_number", "pan_number", "esic_number"]
#     ]
#     for col_idx, value in enumerate(compliance_data, start=19):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     # Compensation Details
#     try:
#         compensation = EmployeeCTC.objects.get(applicant=applicant)
#         compensation_data = [
#             compensation.salary.gross_salary(),
#             compensation.contribution.emp_contribution(),
#             compensation.deduction.emp_deduction(),
#             compensation.calculate_ctc()
#         ]
#     except EmployeeCTC.DoesNotExist:
#         compensation_data = ["N/A"] * 4

#     for col_idx, value in enumerate(compensation_data, start=22):
#         sheet.cell(row=row_idx, column=col_idx, value=value)

#     return next_row


# # Write data for a section
# def write_section_data(sheet, start_row, items, start_col, data_extractor):
#     if items.exists():
#         for item in items:
#             data = data_extractor(item)
#             for col_idx, value in enumerate(data, start=start_col):
#                 sheet.cell(row=start_row, column=col_idx, value=value)
#             start_row += 1
#     else:
#         for col_idx in range(start_col, start_col + len(data_extractor(items.first()))):
#             sheet.cell(row=start_row, column=col_idx, value="N/A")
#         start_row += 1

#     return start_row


# # Export applicants to Excel
# def export_applicants_to_excel(request):
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Applicants Data"

#     create_headers(sheet)

#     applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
#         'profile__educational_qualifications',
#         'profile__employment_records',
#         'profile__languages',
#         'profile__references'
#     )

#     row_idx = 3
#     for applicant in applicants:
#         try:
#             row_idx = write_applicant_data(sheet, row_idx, applicant)
#         except Exception as e:
#             print(f"Error processing applicant {applicant}: {e}")
#         row_idx += 1

#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
#     workbook.save(response)
#     return response

    
##############################################


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from applicant.models import Applicant
from users.models import Profile
from compensation.models import EmployeeCTC

# Create headers for the Excel sheet
def create_headers(sheet):
    sections = [
        ("Educational Qualifications", 1, 5),
        ("Employment Records", 6, 10),
        ("Language Proficiencies", 11, 14),
        ("References", 15, 18),
        ("Compliance Information", 19, 21),
        ("Compensation Details", 22, 25),
    ]

    # Merge cells for section titles
    for title, start_col, end_col in sections:
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = sheet.cell(row=1, column=start_col, value=title)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Define headers for all columns
    headers = [
        # Educational Qualifications
        "Examination Passed", "Year of Passing", "School/College", "Subjects", "Division",
        # Employment Records
        "Organization", "Designation", "Joining Date", "Leaving Date", "Document",
        # Language Proficiencies
        "Language", "Speak", "Read", "Write",
        # References
        "Name", "Occupation", "Phone Number", "Email",
        # Compliance Information
        "UAN Number", "PAN Number", "ESIC Number",
        # Compensation Details
        "Gross Salary", "Total Contribution", "Total Deductions", "Net CTC",
    ]
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=col_idx, value=header)

def write_section_data(sheet, start_row, items, start_col, data_extractor):
    """
    Writes section data into the Excel sheet.
    - `items`: QuerySet or list of objects to write.
    - `data_extractor`: Function to extract the row's data.
    """
    if items.exists():
        for item in items:
            data = data_extractor(item)
            for col_idx, value in enumerate(data, start=start_col):
                sheet.cell(row=start_row, column=col_idx, value=value)
            start_row += 1  # Increment only when data is written
    return start_row  # Return the row index (unchanged if no data was written)



def write_applicant_data(sheet, start_row, applicant):
    # profile = applicant.profile

    try:
        profile = applicant.profile
    except Profile.DoesNotExist:
        profile = None

    # Prepare all data for the applicant
    qualifications = list(profile.educational_qualifications.all())
    employment_records = list(profile.employment_records.all())
    languages = list(profile.languages.all())
    references = list(profile.references.all())

    # Find the maximum number of rows needed for this applicant
    max_rows = max(
        len(qualifications),
        len(employment_records),
        len(languages),
        len(references),
        1,  # At least one row for compliance and compensation data
    )

    for i in range(max_rows):
        row_data = []

        # Add educational qualifications (or empty)
        if i < len(qualifications):
            q = qualifications[i]
            row_data.extend([q.examination_passed, q.year_of_passing, q.school_or_college, q.subjects, q.division])
        else:
            row_data.extend([""] * 5)

        # Add employment records (or empty)
        if i < len(employment_records):
            er = employment_records[i]
            row_data.extend([er.organization, er.designation, er.joining_date, er.leaving_date, er.document or "N/A"])
        else:
            row_data.extend([""] * 5)

        # Add language proficiencies (or empty)
        if i < len(languages):
            lp = languages[i]
            row_data.extend([lp.language, lp.speak, lp.read, lp.write])
        else:
            row_data.extend([""] * 4)

        # Add references (or empty)
        if i < len(references):
            ref = references[i]
            row_data.extend([ref.name, ref.occupation, ref.phone_number, ref.email])
        else:
            row_data.extend([""] * 4)

        # Compliance information (only once)
        if i == 0:
            compliance_data = [
                getattr(applicant.employee, attr, "N/A")
                for attr in ["uan_number", "pan_number", "esic_number"]
            ]
            row_data.extend(compliance_data)
        else:
            row_data.extend([""] * 3)

        # Compensation details (only once)
        if i == 0:
            try:
                compensation = EmployeeCTC.objects.get(applicant=applicant)
                compensation_data = [
                    compensation.salary.gross_salary(),
                    compensation.contribution.emp_contribution(),
                    compensation.deduction.emp_deduction(),
                    compensation.calculate_ctc()
                ]
            except EmployeeCTC.DoesNotExist:
                compensation_data = ["N/A"] * 4
            row_data.extend(compensation_data)
        else:
            row_data.extend([""] * 4)

        # Write the row data
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=start_row, column=col_idx, value=value)

        # Move to the next row
        start_row += 1

    return start_row


def export_applicants_to_excel(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Applicants Data"

    create_headers(sheet)

    applicants = Applicant.objects.select_related('profile', 'user', 'employee').prefetch_related(
        'profile__educational_qualifications',
        'profile__employment_records',
        'profile__languages',
        'profile__references',
    )

    row_idx = 3  # Start below headers
    for applicant in applicants:
        try:
            row_idx = write_applicant_data(sheet, row_idx, applicant)
        except Exception as e:
            print(f"Error processing applicant {applicant}: {e}")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="applicants_data.xlsx"'
    workbook.save(response)
    return response
